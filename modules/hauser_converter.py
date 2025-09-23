"""
하우저 변환기 모듈
기존 howser_order_converter2.py의 핵심 로직을 웹용으로 변환
"""

import re
import os

class HauserConverter:
    def __init__(self):
        self.reference_data = []
        self.load_reference_data()
    
    def load_reference_data(self):
        """참조 엑셀 파일을 로드"""
        try:
            import openpyxl
            ref_path = os.path.join(os.path.dirname(__file__), '..', '하우저 양식 변환.xlsx')
            if os.path.exists(ref_path):
                ref_wb = openpyxl.load_workbook(ref_path, data_only=True)
                if len(ref_wb.worksheets) >= 3:
                    ref_ws = ref_wb.worksheets[2]  # 3번째 시트 사용
                    for row in ref_ws.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1] and row[2] and row[3]:  # 모든 값이 있는 경우만
                            self.reference_data.append({
                                'model': row[0],
                                'color': row[1],
                                'height': row[2],
                                'number': row[3],
                            })
        except Exception as e:
            print(f"참조 파일 로드 오류: {str(e)}")
    
    def extract_text_in_parentheses(self, text):
        """괄호 안의 텍스트 추출"""
        match = re.search(r'\((.*?)\)', text)
        return match.group(1) if match else ''
    
    def extract_parentheses(self, text):
        """모든 괄호 안의 텍스트 추출"""
        return [m.strip() for m in re.findall(r'\((.*?)\)', text)] if text else []
    
    def find_hauser_number(self, product_name):
        """상품명에 맞는 하우저 번호 찾기"""
        prod_matches = self.extract_parentheses(product_name)
        
        for ref in self.reference_data:
            ref_model = str(ref['model']).strip() if ref['model'] else ''
            ref_color = str(ref['color']).strip() if ref['color'] else ''
            ref_height = str(ref['height']).strip() if ref['height'] else ''
            
            # 모델명: 상품명 괄호 안에 참조 모델명이 있으면 OK
            model_match = ref_model in prod_matches
            # 색상: 포함 여부
            color_match = ref_color in product_name
            # 높이: 상품명 괄호 안에 참조 높이가 있으면 OK
            height_match = ref_height in prod_matches
            
            if model_match and color_match and height_match:
                return ref['number']
        
        return None
    
    def convert_excel_file(self, excel_file_path):
        """
        엑셀 파일을 변환 (원본 프로그램과 동일한 로직)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font
            
            # 원본 파일 로드
            wb = openpyxl.load_workbook(excel_file_path)
            ws = wb.active
            
            # 1단계: K열 상품명 읽기 및 줄바꿈 기준 분리
            product_names_per_row = []  # 각 행별 상품명 리스트 저장
            for row in ws.iter_rows(min_row=2):  # 1행은 헤더이므로 2행부터
                k_cell = row[10]  # K열: 0-indexed로 10
                if k_cell.value:
                    # 셀 값이 문자열이면 줄바꿈(\n, \r, \r\n) 기준 분리
                    names = str(k_cell.value).replace('\r\n', '\n').replace('\r', '\n').split('\n')
                    # 빈 문자열 제거
                    names = [n.strip() for n in names if n.strip()]
                else:
                    names = []
                product_names_per_row.append(names)
            
            # 2단계: 상품명 개수만큼 K~L열 복사 및 붙여넣기, 번호 붙이기, 파란색 처리
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                names = product_names_per_row[idx-2]
                if not names:
                    continue
                # K~L열 값 복사
                k_cell = ws.cell(row=idx, column=11)  # K열
                l_cell = ws.cell(row=idx, column=12)  # L열
                gray_font = Font(color='888888')  # 회색 강조3
                k_cell.font = gray_font
                l_cell.font = gray_font
                k_value = row[10].value  # 상품명
                l_value = row[11].value  # 수량
                # 상품명 개수와 수량 비교
                try:
                    qty = int(l_value)
                except (TypeError, ValueError):
                    qty = 0
                if len(names) == 1:
                    # 상품이 1개면 그대로 M, N열에 복사
                    ws.cell(row=idx, column=13, value=k_value)  # M열
                    ws.cell(row=idx, column=14, value=l_value)  # N열
                else:
                    for i, name in enumerate(names):
                        col_offset = 12 + i*2  # M=13, O=15, Q=17 ...
                        new_name = f"{name}{i+1}" if len(names) > 1 else name
                        ws.cell(row=idx, column=col_offset+1, value=new_name)  # 상품명
                        # 조건에 따라 수량/색상 처리
                        if len(names) == qty and qty > 0:
                            ws.cell(row=idx, column=col_offset+2, value=1)  # 수량 1 입력
                            ws.cell(row=idx, column=col_offset+1).font = Font(color='000000')  # 검정
                            ws.cell(row=idx, column=col_offset+2).font = Font(color='000000')  # 검정
                        else:
                            ws.cell(row=idx, column=col_offset+2, value=None)  # 빈칸
                            ws.cell(row=idx, column=col_offset+1).font = Font(color='0000FF')  # 파랑
                            ws.cell(row=idx, column=col_offset+2).font = Font(color='0000FF')  # 파랑

            # 3단계: 상품명 → 하우저번호 변환, 변환 불가시 빨간색 처리
            invalid_rows = []  # 변환이 안 된 행 인덱스 저장
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                names = product_names_per_row[idx-2]
                if not names:
                    continue
                row_invalid = False
                for i, name in enumerate(names):
                    col_offset = 12 + i*2  # M=13, O=15, Q=17 ...
                    cell = ws.cell(row=idx, column=col_offset+1)
                    if cell.value:
                        hauser_num = self.find_hauser_number(cell.value)
                        if hauser_num:
                            cell.value = hauser_num
                        else:
                            # 변환 불가: 빨간색 처리
                            cell.font = Font(color='FF0000')
                            row_invalid = True
                if row_invalid:
                    invalid_rows.append(idx)

            # 변환이 안 된 행 전체 빨간색 처리 및 하단 이동
            if invalid_rows:
                # 행 데이터 복사
                rows_to_move = []
                for idx in invalid_rows:
                    row_data = [ws.cell(row=idx, column=col).value for col in range(1, ws.max_column+1)]
                    rows_to_move.append(row_data)
                # 행을 역순으로 삭제(위치가 밀리지 않게)
                for idx in sorted(invalid_rows, reverse=True):
                    ws.delete_rows(idx)
                # 마지막 데이터 행 이후의 빈 행을 모두 삭제
                while ws.max_row > 1 and all(ws.cell(row=ws.max_row, column=col).value in (None, '') for col in range(1, ws.max_column+1)):
                    ws.delete_rows(ws.max_row)
                # 맨 아래에 붙여넣기 + 빨간색 처리
                for row_data in rows_to_move:
                    ws.append(row_data)
                    last_row = ws.max_row
                    for col in range(1, ws.max_column+1):
                        ws.cell(row=last_row, column=col).font = Font(color='FF0000')
            
            # 수정된 파일 저장
            wb.save(excel_file_path)
            return {'success': True, 'message': '변환 완료'}
            
        except Exception as e:
            return {'error': f"엑셀 파일 처리 오류: {str(e)}"}
    
    def create_result_excel(self, result_data, output_path):
        """
        결과를 엑셀 파일로 저장
        
        Args:
            result_data (list): 변환된 데이터
            output_path (str): 저장할 파일 경로
        
        Returns:
            bool: 성공 여부
        """
        try:
            import openpyxl
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "변환 결과"
            
            # 헤더 추가
            headers = ['상품명', '하우저 번호', '원본 행 번호']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=1, column=col).font = Font(bold=True)
            
            # 데이터 추가
            for row, data in enumerate(result_data, 2):
                if 'error' not in data:
                    ws.cell(row=row, column=1, value=data.get('product_name', ''))
                    ws.cell(row=row, column=2, value=data.get('hauser_number', ''))
                    ws.cell(row=row, column=3, value=data.get('row_number', ''))
                else:
                    ws.cell(row=row, column=1, value=data['error'])
                    ws.cell(row=row, column=1).font = Font(color='FF0000')  # 빨간색
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"결과 파일 저장 오류: {str(e)}")
            return False
    
    def get_reference_info(self):
        """참조 데이터 정보 반환"""
        return {
            'total_count': len(self.reference_data),
            'sample_data': self.reference_data[:5] if self.reference_data else []
        }




